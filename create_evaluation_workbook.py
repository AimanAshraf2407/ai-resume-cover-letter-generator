import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ==========================================
# LOAD TEST DATASET
# ==========================================

input_file = "tests/test_dataset.csv"

df = pd.read_csv(input_file)


# ==========================================
# CREATE WORKBOOK
# ==========================================

output_file = "results/evaluation_results.xlsx"

workbook = Workbook()


# ==========================================
# SHEET 1 - TEST CASES
# ==========================================

ws = workbook.active
ws.title = "Test Cases"

test_columns = [
    "Test ID",
    "Scenario",
    "Source ID",
    "Category",
    "Resume Text",
    "Job Description",
    "Required Skills",
    "Resume Skills",
    "Baseline AI Match Score",
    "Baseline Skill Match Score",
    "Baseline Fuzzy Match Score"
]

ws.append(test_columns)

for _, row in df.iterrows():

    ws.append([
        row["test_id"],
        row["scenario"],
        row["source_id"],
        row["category"],
        row["resume_text"],
        row["job_text"],
        row["job_required_skills"],
        row["resume_skill_list"],
        row["baseline_ai_match_score"],
        row["baseline_skill_match_score"],
        row["baseline_fuzzy_match_score"]
    ])


# ==========================================
# SHEET 2 - RESUME EVALUATION
# ==========================================

ws_resume = workbook.create_sheet("Resume Evaluation")

resume_columns = [
    "Test ID",
    "Information Preservation /5",
    "Completeness /5",
    "Accuracy /5",
    "Job Relevance /5",
    "Professional Quality /5",
    "Hallucination Detected?",
    "Comments",
    "Total Score /25"
]

ws_resume.append(resume_columns)

for test_id in df["test_id"]:

    ws_resume.append([
        test_id,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        f'=SUM(B{ws_resume.max_row + 1}:F{ws_resume.max_row + 1})'
    ])


# ==========================================
# SHEET 3 - COVER LETTER EVALUATION
# ==========================================

ws_cover = workbook.create_sheet("Cover Letter Evaluation")

cover_columns = [
    "Test ID",
    "Personalization /5",
    "Job Relevance /5",
    "Accuracy /5",
    "Professional Tone /5",
    "Completeness /5",
    "Hallucination Detected?",
    "Comments",
    "Total Score /25"
]

ws_cover.append(cover_columns)

for test_id in df["test_id"]:

    ws_cover.append([
        test_id,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        f'=SUM(B{ws_cover.max_row + 1}:F{ws_cover.max_row + 1})'
    ])


# ==========================================
# SHEET 4 - JOB MATCHING
# ==========================================

ws_match = workbook.create_sheet("Job Matching")

match_columns = [
    "Test ID",
    "Required Skills",
    "AI Matched Skills",
    "Correct Matches",
    "Missing Skills",
    "False Matches",
    "AI Match Score /100",
    "Comments"
]

ws_match.append(match_columns)

for _, row in df.iterrows():

    ws_match.append([
        row["test_id"],
        row["job_required_skills"],
        row.get("baseline_ai_match_score", ""),
        "",
        "",
        "",
        "",
        ""
    ])


# ==========================================
# SHEET 5 - SUMMARY
# ==========================================

ws_summary = workbook.create_sheet("Summary")

summary_data = [
    ["AI RESUME & COVER LETTER GENERATOR"],
    ["Testing and Evaluation Summary"],
    [""],
    ["Metric", "Result"],
    ["Number of Test Cases", len(df)],
    ["Average Resume Score /25", '=AVERAGE(\'Resume Evaluation\'!I2:I21)'],
    ["Average Cover Letter Score /25", '=AVERAGE(\'Cover Letter Evaluation\'!I2:I21)'],
    ["Average Resume Score /100", '=AVERAGE(\'Resume Evaluation\'!I2:I21)*4'],
    ["Average Cover Letter Score /100", '=AVERAGE(\'Cover Letter Evaluation\'!I2:I21)*4'],
    ["Hallucination Cases", '=COUNTIF(\'Resume Evaluation\'!G2:G21,"Yes")+COUNTIF(\'Cover Letter Evaluation\'!G2:G21,"Yes")'],
    ["Total Evaluation Scores", 40],
]

for row in summary_data:
    ws_summary.append(row)


# ==========================================
# SHEET 6 - EVALUATION GUIDE
# ==========================================

ws_guide = workbook.create_sheet("Evaluation Guide")

guide_data = [
    ["EVALUATION CRITERIA"],
    [""],
    ["RESUME EVALUATION"],
    ["Information Preservation"],
    ["5 = All important information from the input is preserved"],
    ["4 = Almost all important information is preserved"],
    ["3 = Some information is missing"],
    ["2 = Many important details are missing"],
    ["1 = Very little important information is preserved"],
    [""],
    ["Completeness"],
    ["5 = Complete resume with all relevant sections"],
    ["4 = Minor information missing"],
    ["3 = Some sections incomplete"],
    ["2 = Several important sections missing"],
    ["1 = Very incomplete"],
    [""],
    ["Accuracy"],
    ["5 = No factual errors"],
    ["4 = Minor errors"],
    ["3 = Some noticeable errors"],
    ["2 = Several errors"],
    ["1 = Major factual errors"],
    [""],
    ["Job Relevance"],
    ["5 = Highly tailored to the job description"],
    ["4 = Mostly relevant"],
    ["3 = Moderately relevant"],
    ["2 = Limited relevance"],
    ["1 = Not relevant"],
    [""],
    ["Professional Quality"],
    ["5 = Excellent professional structure and language"],
    ["4 = Good professional quality"],
    ["3 = Acceptable"],
    ["2 = Needs improvement"],
    ["1 = Poor quality"],
    [""],
    ["HALLUCINATION"],
    ["Yes = AI invented information not provided by the user"],
    ["No = AI did not invent information"],
    [""],
    ["COVER LETTER EVALUATION"],
    ["Personalization = How well the letter is customized to the candidate and job"],
    ["Job Relevance = How well it addresses the target position"],
    ["Accuracy = Whether facts about the candidate are correct"],
    ["Professional Tone = Professional writing and appropriate language"],
    ["Completeness = Contains the expected cover letter components"],
]

for row in guide_data:
    ws_guide.append(row)


# ==========================================
# FORMATTING
# ==========================================

for worksheet in workbook.worksheets:

    # Make header row bold
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Wrap text
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # Set reasonable column widths
    for column_cells in worksheet.columns:

        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells:

            if cell.value is not None:

                value_length = len(
                    str(cell.value)
                )

                if value_length > max_length:
                    max_length = value_length

        worksheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 50)


# ==========================================
# SAVE WORKBOOK
# ==========================================

workbook.save(output_file)

print("==========================================")
print("EVALUATION WORKBOOK CREATED")
print("==========================================")
print("Test cases:", len(df))
print("Output file:")
print(output_file)
print("==========================================")